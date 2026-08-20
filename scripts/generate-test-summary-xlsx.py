#!/usr/bin/env python3
"""Generate main-report/test-summary.xlsx for HW06 submission."""
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "main-report" / "test-summary.xlsx"

SUMMARY = [
    ("Student ID", "23127273"),
    ("APIs tested", "FR06, FR07, FR16"),
    ("Tool", "Postman + Newman 6.2.2"),
    ("SUT", "http://127.0.0.1:3000"),
    ("", ""),
    ("Metric", "FR06", "FR07", "FR16", "Total"),
    ("Generated (AI)", 40, 40, 40, 120),
    ("Audited VALID", 32, 40, 40, 112),
    ("Extended (manual)", 6, 6, 6, 18),
    ("Executed (CSV rows)", 48, 46, 46, 140),
    ("Assertions passed", 184, 164, 163, 511),
    ("Assertions failed", 24, 26, 23, 73),
    ("Bugs found", 6, 3, 6, 15),
    ("GitHub Issues", "#1-#6", "#7-#9", "#10-#15", "#1-#15"),
]

APIS = [
    ("FR06 Product Detail", "fr06-product-detail-data.csv"),
    ("FR07 Shopping Cart", "fr07-shopping-cart-data.csv"),
    ("FR16 Product Import", "fr16-product-import-data.csv"),
]


def style_header(ws, row, cols):
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def add_summary(wb):
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 28
    for col in "BCDE":
        ws.column_dimensions[col].width = 16
    for r, row in enumerate(SUMMARY, 1):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    ws.cell(row=6, column=1).font = Font(bold=True)
    for c in range(1, 6):
        ws.cell(row=6, column=c).font = Font(bold=True)


def add_api_sheet(wb, title, csv_name):
    path = ROOT / "postman" / "data" / csv_name
    with path.open(newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    ws = wb.create_sheet(title[:31])
    headers = list(rows[0].keys()) if rows else []
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))
    for r, row in enumerate(rows, 2):
        for c, h in enumerate(headers, 1):
            ws.cell(row=r, column=c, value=row.get(h, ""))
    for col in ws.columns:
        letter = col[0].column_letter
        ws.column_dimensions[letter].width = min(40, max(12, len(str(col[0].value or "")) + 2))


def main():
    wb = Workbook()
    add_summary(wb)
    for title, csv_name in APIS:
        add_api_sheet(wb, title, csv_name)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
